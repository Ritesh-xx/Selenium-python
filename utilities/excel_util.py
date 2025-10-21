import openpyxl

class ExcelUtil:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook[self.sheet_name]

    def get_row_count(self):
        return self.sheet.max_row

    def get_column_count(self):
        return self.sheet.max_column

    def get_cell_data(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def get_data_as_list(self):
        data = []
        for row in range(2, self.get_row_count() + 1):  # assuming first row is header
            row_data = []
            for col in range(1, self.get_column_count() + 1):
                row_data.append(self.get_cell_data(row, col))
            data.append(row_data)
        return data
